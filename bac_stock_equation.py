import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
wb = openpyxl.load_workbook('bac.xlsx')
type(wb)
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet1')
print sheet
anotherSheet = wb.active
print anotherSheet

origin = 0
time = 0
timeaxis = []
priceaxis = []
equ_axis = []

parameter = [0.00112458983045,0.00112458983045,-8.00882016243e-05]

ALPHA = 0.000000000008

for i in reversed(range(1,128,1)):
	date = anotherSheet.cell(row=i, column=1).value
	close_price = anotherSheet.cell(row=i, column=3).value
	date_array = str(date).split('-')
	month_iter = int(date_array[1])
	calendar = {1:31, 2:28, 3:31, 4:30, 5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}
	while(month_iter!=0):
		month_iter-=1
		time+=calendar.get(month_iter,0)
	time+=int(date_array[2].split()[0])
	if(origin==0):
		origin=time
	time-=origin
	time=time-64/128 #mean normalization and feature scaling
	timeaxis.append(time)
	priceaxis.append(close_price)
	time=0
nptime = np.array(timeaxis)
print "It worked!1"

iter = 0
while iter<1000000: #how many iterations you want the program to run for. More accurate the more iterations
	sum_one = 0
	sum_two = 0
	sum_three = 0
	hypo_sum = 0
	for x in range(0,len(parameter)):
		hypo_sum += np.power(nptime, x)*parameter[x]

	for x in range(0,len(parameter)):
		parameter[x] -= ALPHA * np.sum(np.multiply(hypo_sum - np.array(priceaxis),np.power(nptime, x)))

	#print "This is the index of the iteration", iter, "theta_first:", parameter[0], "theta_second:", parameter[1], "theta_third", parameter[2]
	iter+=1
print "This is the index of the iteration", iter, "theta_first:", parameter[0], "theta_second:", parameter[1], "theta_third", parameter[2]
for i in xrange(0,127,1):
	x = timeaxis[i]
	equ_axis.append(parameter[0] + parameter[1]*x + parameter[2]*x*x)
print "equation points plotted!"

plt.plot(timeaxis,priceaxis)
plt.savefig("actualdata.png")
plt.plot(timeaxis,equ_axis)
plt.savefig("equationed.png")